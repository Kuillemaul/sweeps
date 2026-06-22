from __future__ import annotations

import csv
import json
import math
import random
import re
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
from urllib.parse import urlparse


NORMAL_PRICE_CENTS = 50
CUP_SWEEPS = [("Cup $1", 100), ("Cup $2", 200), ("Cup $5", 500)]

DEFAULT_PAYOUTS_BY_POOL_CENTS = {
    350: (200, 100, 50),
    400: (250, 100, 50),
    500: (300, 150, 50),
    600: (300, 200, 100),
    650: (350, 200, 100),
    750: (450, 200, 100),
    800: (500, 200, 100),
    2300: (1300, 800, 200),
    4600: (2600, 1600, 400),
    11500: (7000, 3500, 1000),
}

# Cup defaults reflect Brad's previous payout table. They intentionally do not
# blindly follow percentage splits, because the office sweep uses friendly cash
# amounts rather than exact proportional payouts.
DEFAULT_PAYOUTS_BY_SWEEP_LABEL = {
    "Cup $1": (1300, 800, 200),
    "Cup $2": (2600, 1600, 400),
    "Cup $5": (7000, 3500, 1000),
}


@dataclass
class Attendee:
    attendee_id: str
    name: str
    active: bool = True
    cup_eligible: bool = True
    paid: bool = False


@dataclass
class Horse:
    race_number: int
    horse_number: int
    horse_name: str
    barrier: str = ""
    jockey: str = ""
    trainer: str = ""
    weight: str = ""
    result_position: Optional[int] = None
    status: str = "Starter"
    odds: str = ""

    @property
    def is_runner(self) -> bool:
        status = (self.status or "").strip().lower()
        return status not in {"scratched", "non-runner", "non runner", "withdrawn"}


@dataclass
class Race:
    race_number: int
    race_name: str
    start_time: str = ""
    distance: str = ""
    horses: List[Horse] = field(default_factory=list)

    @property
    def runners(self) -> List[Horse]:
        return [h for h in self.horses if h.is_runner]

    def set_results(self, first: int, second: int, third: int) -> None:
        seen = {first, second, third}
        if len(seen) != 3:
            raise ValueError("1st, 2nd and 3rd horse numbers must be different.")
        for horse in self.horses:
            if horse.horse_number == first:
                horse.result_position = 1
            elif horse.horse_number == second:
                horse.result_position = 2
            elif horse.horse_number == third:
                horse.result_position = 3
            elif horse.result_position in {1, 2, 3}:
                horse.result_position = None


@dataclass
class Allocation:
    race_number: int
    race_name: str
    sweep_label: str
    sweep_round: int
    price_cents: int
    horse_number: int
    horse_name: str
    attendee_id: str
    attendee_name: str
    allocation_type: str
    barrier: str = ""
    jockey: str = ""
    trainer: str = ""
    result_position: Optional[int] = None
    odds: str = ""


@dataclass
class PayoutRow:
    race_number: int
    race_name: str
    sweep_label: str
    placing: int
    horse_number: int
    horse_name: str
    attendee_name: str
    payout_cents: int
    eligible: bool
    note: str


@dataclass
class AuditEntry:
    timestamp: str
    action: str
    details: str


class SweepBook:
    def __init__(self) -> None:
        self.attendees: List[Attendee] = []
        self.races: Dict[int, Race] = {}
        self.allocations: List[Allocation] = []
        self._extra_used_cycle: set[str] = set()
        self.random_seed: Optional[int] = None
        self.locked_sweeps: set[Tuple[int, str]] = set()
        self.audit_log: List[AuditEntry] = []
        self.payout_settings: Dict[str, Tuple[int, int, int]] = default_payout_settings()

    @property
    def active_attendees(self) -> List[Attendee]:
        return [a for a in self.attendees if a.active]

    @property
    def cup_attendees(self) -> List[Attendee]:
        return [a for a in self.attendees if a.active and a.cup_eligible]

    def audit(self, action: str, details: str) -> None:
        self.audit_log.append(AuditEntry(datetime.now().strftime("%Y-%m-%d %H:%M:%S"), action, details))

    def set_seed(self, seed: Optional[int]) -> None:
        self.random_seed = seed

    def rng(self, salt: int = 0) -> random.Random:
        if self.random_seed is None:
            return random.Random()
        return random.Random(self.random_seed + salt)

    def lock_sweep(self, race_number: int, sweep_label: str) -> None:
        self.locked_sweeps.add((race_number, sweep_label))
        self.audit("Locked sweep", f"Race {race_number} - {sweep_label}")

    def unlock_sweep(self, race_number: int, sweep_label: str) -> None:
        self.locked_sweeps.discard((race_number, sweep_label))
        self.audit("Unlocked sweep", f"Race {race_number} - {sweep_label}")

    def is_sweep_locked(self, race_number: int, sweep_label: str) -> bool:
        return (race_number, sweep_label) in self.locked_sweeps

    def is_race_locked(self, race_number: int) -> bool:
        return any(race_no == race_number for race_no, _label in self.locked_sweeps)

    def locked_labels_for_race(self, race_number: int) -> List[str]:
        return sorted(label for race_no, label in self.locked_sweeps if race_no == race_number)

    def _raise_if_race_locked(self, race_number: int, action: str) -> None:
        labels = self.locked_labels_for_race(race_number)
        if labels:
            raise ValueError(f"Cannot {action}; Race {race_number} has locked sweep(s): {', '.join(labels)}.")

    def clear_allocations_for_race(self, race_number: int, *, force: bool = False) -> None:
        if not force:
            self._raise_if_race_locked(race_number, "clear allocations")
        self.allocations = [a for a in self.allocations if a.race_number != race_number]
        if force:
            self.locked_sweeps = {key for key in self.locked_sweeps if key[0] != race_number}

    def clear_cup_allocations(self, *, force: bool = False) -> None:
        if not force:
            self._raise_if_race_locked(7, "clear Cup allocations")
        self.allocations = [a for a in self.allocations if not (a.race_number == 7 and a.sweep_label.startswith("Cup"))]
        if force:
            self.locked_sweeps = {key for key in self.locked_sweeps if key[0] != 7}

    def clear_cup_sweep(self, sweep_label: str, *, force: bool = False) -> None:
        if self.is_sweep_locked(7, sweep_label) and not force:
            raise ValueError(f"Cannot clear {sweep_label}; it is locked.")
        self.allocations = [a for a in self.allocations if not (a.race_number == 7 and a.sweep_label == sweep_label)]
        if force:
            self.locked_sweeps.discard((7, sweep_label))

    def attendee_by_id(self, attendee_id: str) -> Optional[Attendee]:
        return next((a for a in self.attendees if a.attendee_id == attendee_id), None)

    def replace_race(self, race: Race) -> None:
        self._raise_if_race_locked(race.race_number, "replace race data")
        self.races[race.race_number] = race
        self.clear_allocations_for_race(race.race_number)
        self.audit("Imported race", f"Race {race.race_number}: {race.race_name} ({len(race.runners)} runners)")

    def replace_attendees(self, attendees: List[Attendee]) -> None:
        if self.locked_sweeps:
            raise ValueError("Cannot replace attendees while one or more sweeps are locked. Unlock sweeps or start a new event first.")
        self.attendees = attendees
        # Existing allocations may point at old attendee IDs, so reset the generated draw data.
        self.allocations = []
        self._extra_used_cycle.clear()
        self.audit("Imported attendees", f"{len(attendees)} attendees loaded; allocations cleared")

    def _next_extra_attendee(self, eligible: List[Attendee], rng: random.Random) -> Attendee:
        if not eligible:
            raise ValueError("No eligible attendees available for extra fill-in allocation.")
        eligible_ids = {a.attendee_id for a in eligible}
        available = [a for a in eligible if a.attendee_id not in self._extra_used_cycle]
        if not available:
            self._extra_used_cycle = {x for x in self._extra_used_cycle if x not in eligible_ids}
            available = eligible[:]
        chosen = rng.choice(available)
        self._extra_used_cycle.add(chosen.attendee_id)
        return chosen

    def generate_normal_race(self, race_number: int) -> List[Allocation]:
        if race_number == 7:
            return self.generate_cup_sweeps()
        self._raise_if_race_locked(race_number, "regenerate race")
        race = self.races.get(race_number)
        if not race:
            raise ValueError(f"Race {race_number} was not found.")
        attendees = self.active_attendees
        horses = race.runners
        if not attendees:
            raise ValueError("No active attendees found.")
        if not horses:
            raise ValueError(f"Race {race_number} has no runners.")
        self.clear_allocations_for_race(race_number)
        rng = self.rng(race_number * 1000)
        sweep_count = max(1, math.ceil(len(attendees) / len(horses)))
        standard_pool = attendees[:]
        rng.shuffle(standard_pool)
        new_allocations: List[Allocation] = []
        for sweep_round in range(1, sweep_count + 1):
            horses_this_round = horses[:]
            rng.shuffle(horses_this_round)
            for horse in horses_this_round:
                if standard_pool:
                    attendee = standard_pool.pop(0)
                    allocation_type = "Standard"
                else:
                    attendee = self._next_extra_attendee(attendees, rng)
                    allocation_type = "Extra Fill-In"
                new_allocations.append(
                    Allocation(
                        race_number=race.race_number,
                        race_name=race.race_name,
                        sweep_label=f"Sweep {sweep_round}",
                        sweep_round=sweep_round,
                        price_cents=NORMAL_PRICE_CENTS,
                        horse_number=horse.horse_number,
                        horse_name=horse.horse_name,
                        attendee_id=attendee.attendee_id,
                        attendee_name=attendee.name,
                        allocation_type=allocation_type,
                        barrier=horse.barrier,
                        jockey=horse.jockey,
                        trainer=horse.trainer,
                        result_position=horse.result_position,
                        odds=getattr(horse, "odds", ""),
                    )
                )
        self.allocations.extend(new_allocations)
        self.audit("Generated sweeps", f"Race {race_number}: {len(new_allocations)} allocations across {sweep_count} sweep(s)")
        return new_allocations

    def generate_all_normal_races(self) -> None:
        for race_number in sorted(self.races):
            if race_number != 7:
                self.generate_normal_race(race_number)

    def generate_cup_sweep(self, sweep_label: str) -> List[Allocation]:
        sweep_lookup = dict(CUP_SWEEPS)
        if sweep_label not in sweep_lookup:
            raise ValueError(f"Unknown Cup sweep: {sweep_label}")
        if self.is_sweep_locked(7, sweep_label):
            raise ValueError(f"Cannot regenerate {sweep_label}; it is locked.")
        price_cents = sweep_lookup[sweep_label]
        race = self.races.get(7)
        if not race:
            raise ValueError("Race 7, the Melbourne Cup, was not found.")
        horses = race.runners
        attendees = self.cup_attendees
        if not attendees:
            raise ValueError("No Cup-eligible attendees found.")
        if not horses:
            raise ValueError("Race 7 has no runners.")

        self.clear_cup_sweep(sweep_label)
        sweep_index = [label for label, _price in CUP_SWEEPS].index(sweep_label) + 1
        rng = self.rng(7000 + sweep_index * 100)

        horses_this_sweep = horses[:]
        rng.shuffle(horses_this_sweep)
        attendee_pool = attendees[:]
        rng.shuffle(attendee_pool)
        if len(attendee_pool) > len(horses_this_sweep):
            attendee_pool = attendee_pool[: len(horses_this_sweep)]

        new_allocations: List[Allocation] = []
        for horse in horses_this_sweep:
            if attendee_pool:
                attendee = attendee_pool.pop(0)
                allocation_type = "Standard"
            else:
                attendee = self._next_extra_attendee(attendees, rng)
                allocation_type = "Extra Fill-In"
            new_allocations.append(
                Allocation(
                    race_number=7,
                    race_name=race.race_name,
                    sweep_label=sweep_label,
                    sweep_round=sweep_index,
                    price_cents=price_cents,
                    horse_number=horse.horse_number,
                    horse_name=horse.horse_name,
                    attendee_id=attendee.attendee_id,
                    attendee_name=attendee.name,
                    allocation_type=allocation_type,
                    barrier=horse.barrier,
                    jockey=horse.jockey,
                    trainer=horse.trainer,
                    result_position=horse.result_position,
                    odds=getattr(horse, "odds", ""),
                )
            )
        self.allocations.extend(new_allocations)
        self.audit("Generated Cup sweep", f"{sweep_label}: {len(new_allocations)} allocations")
        return new_allocations

    def generate_cup_sweeps(self) -> List[Allocation]:
        self._raise_if_race_locked(7, "regenerate Cup sweeps")
        self.clear_cup_allocations()
        generated: List[Allocation] = []
        for label, _price_cents in CUP_SWEEPS:
            generated.extend(self.generate_cup_sweep(label))
        return generated

    def amount_owing_rows(self) -> List[Dict[str, object]]:
        rows = []
        for attendee in sorted(self.attendees, key=lambda a: a.name):
            attendee_allocations = [a for a in self.allocations if a.attendee_id == attendee.attendee_id]
            normal = sum(a.price_cents for a in attendee_allocations if a.race_number != 7)
            cup_1 = sum(a.price_cents for a in attendee_allocations if a.sweep_label == "Cup $1")
            cup_2 = sum(a.price_cents for a in attendee_allocations if a.sweep_label == "Cup $2")
            cup_5 = sum(a.price_cents for a in attendee_allocations if a.sweep_label == "Cup $5")
            total = normal + cup_1 + cup_2 + cup_5
            rows.append(
                {
                    "Attendee ID": attendee.attendee_id,
                    "Attendee": attendee.name,
                    "Normal Sweeps": normal,
                    "Cup $1": cup_1,
                    "Cup $2": cup_2,
                    "Cup $5": cup_5,
                    "Total Owing": total,
                    "Paid": attendee.paid,
                }
            )
        return rows

    def payout_amounts_for(self, pool_total_cents: int, sweep_label: str) -> Tuple[int, int, int]:
        label_key = payout_setting_key("label", sweep_label)
        if label_key in self.payout_settings:
            return tuple(self.payout_settings[label_key])  # type: ignore[return-value]
        pool_key = payout_setting_key("pool", str(pool_total_cents))
        if pool_key in self.payout_settings:
            return tuple(self.payout_settings[pool_key])  # type: ignore[return-value]
        return payout_split_default(pool_total_cents)

    def payout_rows(self, race_number: Optional[int] = None) -> List[PayoutRow]:
        allocations = self.allocations
        if race_number is not None:
            allocations = [a for a in allocations if a.race_number == race_number]
        grouped: Dict[Tuple[int, str], List[Allocation]] = {}
        for allocation in allocations:
            grouped.setdefault((allocation.race_number, allocation.sweep_label), []).append(allocation)
        rows: List[PayoutRow] = []
        for (race_no, sweep_label), group in sorted(grouped.items()):
            race = self.races.get(race_no)
            if not race:
                continue
            pool_total = sum(a.price_cents for a in group)
            payout_amounts = self.payout_amounts_for(pool_total, sweep_label)
            winners = {h.result_position: h for h in race.runners if h.result_position in {1, 2, 3}}
            for placing in (1, 2, 3):
                horse = winners.get(placing)
                if not horse:
                    continue
                allocation = next((a for a in group if a.horse_number == horse.horse_number), None)
                if not allocation:
                    continue
                attendee = self.attendee_by_id(allocation.attendee_id)
                eligible = bool(attendee and attendee.paid)
                amount = payout_amounts[placing - 1] if eligible else 0
                note = "Eligible" if eligible else "Blocked - unpaid"
                rows.append(
                    PayoutRow(
                        race_number=race_no,
                        race_name=race.race_name,
                        sweep_label=sweep_label,
                        placing=placing,
                        horse_number=horse.horse_number,
                        horse_name=horse.horse_name,
                        attendee_name=allocation.attendee_name,
                        payout_cents=amount,
                        eligible=eligible,
                        note=note,
                    )
                )
        return rows

    def payout_summary_rows(self) -> List[Dict[str, object]]:
        rows = []
        grouped: Dict[Tuple[int, str], List[Allocation]] = {}
        for allocation in self.allocations:
            grouped.setdefault((allocation.race_number, allocation.sweep_label), []).append(allocation)
        for (race_no, sweep_label), group in sorted(grouped.items()):
            pool_total = sum(a.price_cents for a in group)
            payout = self.payout_amounts_for(pool_total, sweep_label)
            rows.append({
                "Race": race_no,
                "Sweep": sweep_label,
                "Collected": pool_total,
                "Payout Total": sum(payout),
                "Difference": pool_total - sum(payout),
            })
        return rows

    def to_dict(self) -> Dict[str, object]:
        return {
            "version": 2,
            "random_seed": self.random_seed,
            "attendees": [asdict(a) for a in self.attendees],
            "races": [
                {
                    "race_number": race.race_number,
                    "race_name": race.race_name,
                    "start_time": race.start_time,
                    "distance": race.distance,
                    "horses": [asdict(h) for h in race.horses],
                }
                for race in sorted(self.races.values(), key=lambda r: r.race_number)
            ],
            "allocations": [asdict(a) for a in self.allocations],
            "extra_used_cycle": sorted(self._extra_used_cycle),
            "locked_sweeps": [{"race_number": race_no, "sweep_label": label} for race_no, label in sorted(self.locked_sweeps)],
            "payout_settings": {key: list(value) for key, value in sorted(self.payout_settings.items())},
            "audit_log": [asdict(entry) for entry in self.audit_log],
        }

    @classmethod
    def from_dict(cls, data: Dict[str, object]) -> "SweepBook":
        book = cls()
        book.random_seed = parse_int(data.get("random_seed")) if data.get("random_seed") not in (None, "") else None
        book.attendees = [Attendee(**row) for row in data.get("attendees", [])]  # type: ignore[arg-type]
        book.races = {}
        for row in data.get("races", []):  # type: ignore[assignment]
            if not isinstance(row, dict):
                continue
            horses = [Horse(**h) for h in row.get("horses", [])]  # type: ignore[arg-type]
            race = Race(
                race_number=int(row.get("race_number", 0)),
                race_name=str(row.get("race_name", "")),
                start_time=str(row.get("start_time", "") or ""),
                distance=str(row.get("distance", "") or ""),
                horses=horses,
            )
            if race.race_number:
                book.races[race.race_number] = race
        book.allocations = [Allocation(**row) for row in data.get("allocations", [])]  # type: ignore[arg-type]
        book._extra_used_cycle = set(data.get("extra_used_cycle", []))  # type: ignore[arg-type]
        book.locked_sweeps = set()
        for row in data.get("locked_sweeps", []):  # type: ignore[assignment]
            if isinstance(row, dict):
                race_no = parse_int(row.get("race_number"))
                label = str(row.get("sweep_label", ""))
                if race_no is not None and label:
                    book.locked_sweeps.add((race_no, label))
        payout_settings = data.get("payout_settings")
        if isinstance(payout_settings, dict):
            book.payout_settings = {}
            for key, value in payout_settings.items():
                if isinstance(value, (list, tuple)) and len(value) == 3:
                    book.payout_settings[str(key)] = (int(value[0]), int(value[1]), int(value[2]))
        if not book.payout_settings:
            book.payout_settings = default_payout_settings()
        book.audit_log = []
        for row in data.get("audit_log", []):  # type: ignore[assignment]
            if isinstance(row, dict):
                book.audit_log.append(AuditEntry(str(row.get("timestamp", "")), str(row.get("action", "")), str(row.get("details", ""))))
        return book

    def save_event(self, path: str | Path) -> None:
        event_path = Path(path)
        event_path.parent.mkdir(parents=True, exist_ok=True)
        self.audit("Saved event", str(event_path))
        event_path.write_text(json.dumps(self.to_dict(), indent=2), encoding="utf-8")

    @classmethod
    def load_event(cls, path: str | Path) -> "SweepBook":
        event_path = Path(path)
        data = json.loads(event_path.read_text(encoding="utf-8"))
        book = cls.from_dict(data)
        book.audit("Loaded event", str(event_path))
        return book


def payout_setting_key(kind: str, key: str) -> str:
    return f"{kind}:{key}"


def split_payout_setting_key(key: str) -> Tuple[str, str]:
    if ":" not in key:
        return "pool", key
    kind, value = key.split(":", 1)
    return kind, value


def default_payout_settings() -> Dict[str, Tuple[int, int, int]]:
    settings: Dict[str, Tuple[int, int, int]] = {}
    for pool, split in DEFAULT_PAYOUTS_BY_POOL_CENTS.items():
        settings[payout_setting_key("pool", str(pool))] = split
    for label, split in DEFAULT_PAYOUTS_BY_SWEEP_LABEL.items():
        settings[payout_setting_key("label", label)] = split
    return settings


def payout_split_default(pool_total_cents: int) -> Tuple[int, int, int]:
    first = round_to_nearest_50(pool_total_cents * 0.60)
    second = round_to_nearest_50(pool_total_cents * 0.30)
    third = pool_total_cents - first - second
    if third < 0:
        third = 0
    return int(first), int(second), int(third)


def payout_split(pool_total_cents: int) -> Tuple[int, int, int]:
    if pool_total_cents in DEFAULT_PAYOUTS_BY_POOL_CENTS:
        return DEFAULT_PAYOUTS_BY_POOL_CENTS[pool_total_cents]
    return payout_split_default(pool_total_cents)


def round_to_nearest_50(value: float) -> int:
    return int(round(value / 50) * 50)


def money(cents: int) -> str:
    return f"${cents / 100:,.2f}"


def normalise_header(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "_").replace("-", "_")


def first_value(row: Dict[str, object], keys: Iterable[str], default: object = "") -> object:
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return default


def parse_int(value: object, default: Optional[int] = None) -> Optional[int]:
    if value is None or value == "":
        return default
    try:
        text = str(value).strip()
        if text.lower() in {"ur", "unseated", "dnf", "scr", "scratched"}:
            return default
        return int(float(text))
    except Exception:
        return default


def load_workbook_data(path: str | Path) -> SweepBook:
    from openpyxl import load_workbook

    workbook_path = Path(path)
    book = SweepBook()
    if not workbook_path.exists():
        return book
    wb = load_workbook(workbook_path, read_only=True, data_only=True)

    if "Attendees" in wb.sheetnames:
        ws = wb["Attendees"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            headers = [normalise_header(h) for h in rows[0]]
            for index, values in enumerate(rows[1:], start=1):
                row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
                name = first_value(row, ["name", "attendee", "attendee_name"], "")
                if not str(name).strip():
                    continue
                attendee_id = str(first_value(row, ["attendee_id", "id"], f"A{index:03d}")).strip()
                book.attendees.append(
                    Attendee(
                        attendee_id=attendee_id,
                        name=str(name).strip(),
                        active=parse_bool(first_value(row, ["active"], True)),
                        cup_eligible=parse_bool(first_value(row, ["cup_eligible", "cup", "cup_sweep"], True)),
                        paid=parse_bool(first_value(row, ["paid"], False)),
                    )
                )

    race_meta: Dict[int, Dict[str, object]] = {}
    if "Race_Info" in wb.sheetnames:
        ws = wb["Race_Info"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            headers = [normalise_header(h) for h in rows[0]]
            for values in rows[1:]:
                row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
                race_no = parse_int(first_value(row, ["race_number", "race_no", "race", "race_#", "race_id"]))
                if race_no is None:
                    continue
                race_meta[race_no] = row

    runner_sheet_name = next((name for name in ["All_Runners", "Runners"] if name in wb.sheetnames), None)
    if runner_sheet_name:
        ws = wb[runner_sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            headers = [normalise_header(h) for h in rows[0]]
            for values in rows[1:]:
                row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
                race_no = parse_int(first_value(row, ["race_number", "race_no", "race", "race_id", "race_#"]))
                horse_no = parse_int(first_value(row, ["horse_number", "horse_no", "runner_number", "number", "no", "tab_no", "saddlecloth"]))
                horse_name = str(first_value(row, ["horse_name", "runner", "runner_name", "horse", "name"], "")).strip()
                if race_no is None or horse_no is None or not horse_name:
                    continue
                meta = race_meta.get(race_no, {})
                race_name = str(first_value(row, ["race_name", "event", "name_of_race"], first_value(meta, ["race_name", "event", "name"], f"Race {race_no}"))).strip()
                if race_no not in book.races:
                    book.races[race_no] = Race(
                        race_number=race_no,
                        race_name=race_name or f"Race {race_no}",
                        start_time=str(first_value(meta, ["start_time", "time"], "") or ""),
                        distance=str(first_value(meta, ["distance"], "") or ""),
                    )
                status = str(first_value(row, ["status", "runner_status"], "Starter") or "Starter")
                position = parse_int(first_value(row, ["result_position", "position", "placing", "finish_position", "place"]))
                book.races[race_no].horses.append(
                    Horse(
                        race_number=race_no,
                        horse_number=horse_no,
                        horse_name=horse_name,
                        barrier=str(first_value(row, ["barrier", "barrier_number", "gate"], "") or ""),
                        jockey=str(first_value(row, ["jockey", "rider"], "") or ""),
                        trainer=str(first_value(row, ["trainer"], "") or ""),
                        weight=str(first_value(row, ["weight", "weight_kg"], "") or ""),
                        result_position=position,
                        status=status,
                    )
                )
    book.audit("Loaded workbook", str(workbook_path))
    return book


def parse_race_number_from_path(path: str | Path) -> Optional[int]:
    stem = Path(path).stem.lower()
    patterns = [r"(?:^|[-_\s])r(?:ace)?\s*0*(\d{1,2})(?:$|[-_\s])", r"race\s*0*(\d{1,2})"]
    for pattern in patterns:
        match = re.search(pattern, stem)
        if match:
            return parse_int(match.group(1))
    return None


def title_from_slug(slug: str) -> str:
    text = slug.replace("-", " ").replace("_", " ").strip()
    if not text:
        return ""
    small_words = {"and", "of", "the", "to", "for", "in", "on", "at"}
    parts = []
    for index, word in enumerate(text.split()):
        lower = word.lower()
        if index > 0 and lower in small_words:
            parts.append(lower)
        else:
            parts.append(lower.capitalize())
    return " ".join(parts)


def infer_race_name_from_csv_row(row: Dict[str, object], fallback: str) -> str:
    explicit = str(first_value(row, ["race_name", "race", "event", "name_of_race"], "") or "").strip()
    if explicit:
        return explicit
    url = str(first_value(row, ["form_guide_url", "form_url", "url"], "") or "").strip()
    if url:
        try:
            parts = [p for p in urlparse(url).path.split("/") if p]
            if parts:
                race_slug = parts[-1]
                name = title_from_slug(race_slug)
                if name:
                    return name
        except Exception:
            pass
    return fallback


def load_race_csv(path: str | Path, race_number: Optional[int] = None, race_name: str = "") -> Race:
    csv_path = Path(path)
    if not csv_path.exists():
        raise FileNotFoundError(f"Race CSV was not found: {csv_path}")
    parsed_race_number = race_number or parse_race_number_from_path(csv_path) or 1
    horses: List[Horse] = []
    inferred_name = race_name.strip() or f"Race {parsed_race_number}"

    with csv_path.open("r", newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        if not reader.fieldnames:
            raise ValueError("Race CSV has no header row.")
        headers = [normalise_header(h) for h in reader.fieldnames]
        for raw_index, values in enumerate(reader, start=1):
            row = {headers[i]: values.get(reader.fieldnames[i]) for i in range(len(headers))}
            if raw_index == 1 and not race_name.strip():
                inferred_name = infer_race_name_from_csv_row(row, inferred_name)

            horse_no = parse_int(first_value(row, ["num", "number", "horse_no", "horse_number", "runner_number", "saddlecloth", "tab_no"]))
            horse_name = str(first_value(row, ["horse_name", "runner", "runner_name", "horse", "name"], "") or "").strip()
            if horse_no is None or not horse_name:
                continue

            finish_value = first_value(row, ["finish_result_(updates_after_race)", "finish_result", "result", "placing", "place", "finish_position"], "")
            finish_text = str(finish_value or "").strip()
            status = "Starter"
            if finish_text.lower() in {"scr", "scratched", "scratch", "withdrawn", "late scratching", "non-runner", "non runner"}:
                status = "Scratched"
            result_position = parse_int(finish_text)

            horses.append(
                Horse(
                    race_number=parsed_race_number,
                    horse_number=horse_no,
                    horse_name=horse_name,
                    barrier=str(first_value(row, ["barrier", "barrier_number", "gate"], "") or ""),
                    jockey=str(first_value(row, ["jockey", "rider"], "") or ""),
                    trainer=str(first_value(row, ["trainer"], "") or ""),
                    weight=str(first_value(row, ["weight_carried", "weight", "weight_kg"], "") or ""),
                    result_position=result_position,
                    status=status,
                    odds=str(first_value(row, ["best_fixed_odds", "fixed_odds", "odds", "starting_price", "sp", "price", "win_odds", "market_price"], "") or ""),
                )
            )

    if not horses:
        raise ValueError("No runners were found in the race CSV. Expected columns like Num and Horse Name.")

    return Race(
        race_number=parsed_race_number,
        race_name=inferred_name,
        horses=sorted(horses, key=lambda h: h.horse_number),
    )


def load_attendees_file(path: str | Path) -> List[Attendee]:
    attendee_path = Path(path)
    if not attendee_path.exists():
        raise FileNotFoundError(f"Attendees file was not found: {attendee_path}")
    suffix = attendee_path.suffix.lower()
    rows: List[Dict[str, object]] = []

    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        wb = load_workbook(attendee_path, read_only=True, data_only=True)
        ws = wb.active
        values = list(ws.iter_rows(values_only=True))
        if values:
            headers = [normalise_header(h) for h in values[0]]
            # Treat a one-column sheet with no obvious header as a plain list.
            first_header = headers[0] if headers else ""
            if len(headers) == 1 and first_header not in {"name", "attendee", "attendee_name"}:
                rows = [{"name": values[0][0]}] + [{"name": r[0]} for r in values[1:] if r and r[0]]
            else:
                for value_row in values[1:]:
                    rows.append({headers[i]: value_row[i] if i < len(value_row) else None for i in range(len(headers))})
    else:
        text = attendee_path.read_text(encoding="utf-8-sig")
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        if not lines:
            raise ValueError("Attendees file is empty.")
        if "," not in lines[0] and "\t" not in lines[0]:
            rows = [{"name": line} for line in lines]
        else:
            dialect = csv.excel_tab if "\t" in lines[0] and "," not in lines[0] else csv.excel
            reader = csv.DictReader(lines, dialect=dialect)
            if reader.fieldnames:
                normalised = [normalise_header(h) for h in reader.fieldnames]
                if len(normalised) == 1 and normalised[0] not in {"name", "attendee", "attendee_name"}:
                    rows = [{"name": reader.fieldnames[0]}]
                    for raw in reader:
                        rows.append({"name": next(iter(raw.values()), "")})
                else:
                    for raw in reader:
                        rows.append({normalised[i]: raw.get(reader.fieldnames[i]) for i in range(len(normalised))})

    attendees: List[Attendee] = []
    seen: set[str] = set()
    for index, row in enumerate(rows, start=1):
        name = str(first_value(row, ["name", "attendee", "attendee_name", "person"], "") or "").strip()
        if not name:
            continue
        key = name.upper()
        if key in seen:
            continue
        seen.add(key)
        attendee_id = str(first_value(row, ["attendee_id", "id"], f"A{len(attendees) + 1:03d}") or f"A{len(attendees) + 1:03d}").strip()
        attendees.append(
            Attendee(
                attendee_id=attendee_id,
                name=key,
                active=parse_bool(first_value(row, ["active"], True)),
                cup_eligible=parse_bool(first_value(row, ["cup_eligible", "cup", "cup_sweep"], True)),
                paid=parse_bool(first_value(row, ["paid"], False)),
            )
        )

    if not attendees:
        raise ValueError("No attendees were found. Expected a Name/Attendee column or one name per line.")
    return attendees


def parse_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "active", "paid", "eligible"}


def demo_attendees() -> List[Attendee]:
    names = [
        "ERIN", "BRAD", "COLTON", "JENNY", "GREG", "DANIELLE", "LEAH", "STAS", "GRAINNE",
        "JACK", "JAYDE", "MELISSA", "MICHAEL", "MIETTE", "JOANNE", "RAY", "EDDY",
        "SHANNON", "MATT", "IAN", "GLENDA", "HEATHER", "SHIRLEY",
    ]
    return [Attendee(f"A{i:03d}", name, paid=False) for i, name in enumerate(names, start=1)]


def demo_races() -> Dict[int, Race]:
    races: Dict[int, Race] = {}
    for race_no in range(1, 11):
        horse_count = 24 if race_no == 7 else 10 + (race_no % 5)
        race = Race(race_no, "Melbourne Cup" if race_no == 7 else f"Flemington Race {race_no}")
        for horse_no in range(1, horse_count + 1):
            race.horses.append(
                Horse(
                    race_number=race_no,
                    horse_number=horse_no,
                    horse_name=f"Horse {race_no}-{horse_no}",
                    odds=str(2 + horse_no),
                    barrier=str(horse_no),
                    jockey=f"Jockey {horse_no}",
                    trainer=f"Trainer {horse_no}",
                    result_position=horse_no if horse_no <= 3 else None,
                )
            )
        races[race_no] = race
    return races
